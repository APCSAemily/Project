import openpyxl as xl
from openpyxl.chart import BarChart, Reference
from pathlib import Path


def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    cell = sheet['a1']
    # cell = sheet.cell(1,1)
    # print(sheet.max_row)
    sheet['D1'] = 'Corrected Price'

    #range will not include the last value
    for row in range(2,sheet.max_row + 1):
        cell = sheet.cell(row,3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet, 
            min_row=2, 
            max_row=sheet.max_row,
            min_col = 4,
            max_col= 4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')
    wb.save(filename)




if __name__ == "__main__":
    path = Path()
    # Use glob to find all .xlsx files ending with 'xfile'
    files = path.glob('*.xlsx')

    # Process each file found
    for file in files:
        process_workbook(file)
